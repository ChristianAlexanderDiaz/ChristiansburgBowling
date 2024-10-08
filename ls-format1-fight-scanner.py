import pdfplumber
import re
import openpyxl
from datetime import datetime

# Create an exceptions list as a dictionary with NO middle iniitials
name_exceptions = {
    "CHRISTOPHER C. EASTRIDGE": "CHRIS EASTRIDGE",
    "RONNIE DUNCAN": "Ronnie Duncan"
}

def remove_middle_initial(name):
    """
    Removes the middle initial from a name if it exists
    while maintaining Jr., Sr., etc. suffixes.

    Args:
    name (str): The name to process.

    Returns:
    str: The name without the middle initial.
    """
    return re.sub(r"\s+[A-Z]\.\s+", " ", name)

def extract_handicaps(pdf_path):
    """
    Extracts player names, their averages, and corresponding handicaps from the 'Team Rosters' section of a bowling league PDF.

    Args:
    pdf_path (str): The file path to the PDF document.

    Returns:
    list: A list of tuples where each tuple contains (name, average, handicap).
    """
    player_handicaps = []

    with pdfplumber.open(pdf_path) as pdf:
        # Initialize a counter for the pages
        page_number = 1

        # Loop through each page in the PDF
        for page in pdf.pages:
            text = page.extract_text()
            lines = text.splitlines()

            team_roster_found = False

            for line in lines:
                # Stop processing if we reach the 'Temporary Substitutes' section
                if 'Temporary Substitutes' in line:
                    # Sort the list alphabetically by the first name
                    player_handicaps.sort(key=lambda x: x[0])  # Sort by name
                    return player_handicaps

                # Skip team headers like "11 - WYTHEVILLE MOOSE Lane: 22"
                if '-' in line and 'Lane' in line:
                    continue

                # Check for 'Team Rosters' on the first page or 'of' on subsequent pages
                if 'Team Rosters' in line or (page_number > 1 and 'of' in line):
                    team_roster_found = True

                # Once in the 'Team Rosters' section, look for player names and their HDCP
                if team_roster_found:
                    if 'Name' in line and 'Avg HDCP' in line:
                        # Skip header lines, move to the actual player data
                        continue

                    # Look for lines that have player data
                    if any(char.isdigit() for char in line):  # Detect lines with numeric data
                        # Use regex to extract the name, average, and handicap
                        match = re.match(r"([^\d]+)\s+(\d+|bk\d+)\s+(\d+)", line)
                        if match:
                            name = match.group(1).strip()
                            average = match.group(2).strip()
                            handicap = int(match.group(3).strip())
                            
                            # Handle name exceptions if PDF differs from bowling screen
                            if name in name_exceptions:
                                name = name_exceptions[name]
                            else:
                                # Remove the middle initial from the name
                                name = remove_middle_initial(name)
                            player_handicaps.append((name, average, handicap))

            # Increment the page number counter after processing each page
            page_number += 1

    # Sort the list alphabetically by the first name
    player_handicaps.sort(key=lambda x: x[0])  # Sort by name
    return player_handicaps

def update_excel(data, excel_path, sheet_name):
    """
    Updates the player data in the specified Excel sheet and updates F4 with the last update date.

    Args:
    data (list): A list of tuples containing (name, average, handicap).
    excel_path (str): The file path to the Excel document.
    sheet_name (str): The name of the worksheet to update.
    """
    # Load the workbook and select the specified worksheet
    workbook = openpyxl.load_workbook(excel_path, keep_vba=True)
        
    # Print all available sheet names
    print("Available sheets:", workbook.sheetnames)

    # Ensure you use the correct sheet name
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        print(f"Sheet '{sheet_name}' does not exist.")
        # Handle the error as necessary

    # Write the data to the sheet starting at A2
    for i, (name, average, handicap) in enumerate(data, start=2):
        sheet[f"A{i}"] = name
        sheet[f"B{i}"] = average
        sheet[f"C{i}"] = handicap

    # Get the current date and format it
    current_date = datetime.now().strftime("%m/%d %I:%M %p")

    # Update cell F4 with the last updated information
    sheet["F4"] = f"LAST UPDATED: {current_date}"

    # Save the workbook
    workbook.save(excel_path)


# Paths and worksheet name
pdf_path = "/Users/cynical/Documents/GitHub/ChristiansburgBowling/MenWed.pdf"
excel_path = "/Users/cynical/OneDrive/Documents/Wednesday Night Sidepots_MacroCopy.xlsm"  # Synced local path
sheet_name = "Men's Handicap Bank | WEDNESDAY"

# Extract the handicaps
handicaps = extract_handicaps(pdf_path)

# Call the function with your data and worksheet name
update_excel(handicaps, excel_path, sheet_name)

# Optional: Print the results for verification
for entry in handicaps:
    print(entry)
    
print("Reached the end of the script")