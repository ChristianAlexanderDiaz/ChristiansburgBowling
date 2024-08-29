import json
from openpyxl import load_workbook

def update_excel_from_json(json_file, excel_file):
    # Load the JSON data from the file
    with open(json_file, 'r') as f:
        scraped_data = json.load(f)

    # Load the Excel workbook and select the relevant sheet
    workbook = load_workbook(excel_file)
    sheet = workbook["Handicap Sidepot"]

    # Mapping of column indices to game numbers
    game_map = {0: "first", 1: "second", 2: "third"}

    updated_players = []  # Track players who were updated
    active_players = []   # Track players who are still active

    # Iterate over each player in the scraped data
    for player in scraped_data:
        name = player['name']  # Get the player's name
        scores = player['scoresArray']  # Get the array of scores

        # Iterate over the rows in the Excel sheet (rows 2 to 51)
        for row in range(2, 52):
            # Check if the name in the sheet matches the current player's name
            if sheet.cell(row=row, column=2).value == name:
                # Iterate over the scores and update the first empty cell in columns D, E, F
                for i, score in enumerate(scores):
                    cell = sheet.cell(row=row, column=4 + i)  # Calculate the correct column (D, E, F)
                    if cell.value is None:  # Only update the cell if it's empty
                        cell.value = score
                        game = game_map.get(i, "unknown")  # Get the game number (first, second, third)
                        updated_players.append(f"{name} finished their {game} game with a {score}.")
                        break  # Exit the loop once the score is added

                # Check if the player has finished all three games
                if not sheet.cell(row=row, column=6).value:  # Check if the third game is not filled
                    active_players.append(name)

                break  # Exit the loop once the player is found

    # Save the updated Excel file
    workbook.save(excel_file)

    return updated_players, active_players

def main():
    json_file = 'scraped_data.json'
    excel_file = '/Users/cynical/OneDrive/Mario Kart Wii/Documents/Wednesday Night Sidepots_MacroCopy.xlsx'
    
    # Call the update function and retrieve updates and active players
    updated_players, active_players = update_excel_from_json(json_file, excel_file)
    
    print("\n--------------------")
    print("Excel Updates:")
    if updated_players:
        for update in updated_players:
            print(update)
    else:
        print("No updates.")
    print("--------------------")
    
    if active_players:
        print("Active List:")
        for player in active_players:
            print(f"- {player}")
        print("--------------------")

# Usage
main()